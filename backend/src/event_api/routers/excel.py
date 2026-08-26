from __future__ import annotations

import hashlib
import io
import json
from datetime import UTC, date, datetime, timedelta
from typing import Annotated, Any
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, File, Form, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from ..config import Settings
from ..database import Database, execute, row, rows
from ..dependencies import Staff, csrf_super_admin, database, settings, super_admin
from ..errors import ApiError
from ..registration_service import (
    create_registration,
    find_or_create_person,
    participant,
)
from ..schemas import ParticipantValues
from ..service_utils import db_json

router = APIRouter(prefix="/admin/events", tags=["excel"])
MAX_FILE = 5 * 1024 * 1024
MAX_ROWS = 5_000
HEADERS = {
    "lastName": "Фамилия",
    "firstName": "Имя",
    "middleName": "Отчество",
    "birthDate": "Дата рождения",
    "personType": "Тип участника",
    "studyGroup": "Группа",
    "organization": "Организация",
    "phone": "Телефон",
    "email": "Email",
}


def _safe(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def _mapping(headers: list[str], supplied: str | None) -> dict[str, Any]:
    if supplied:
        try:
            result = json.loads(supplied)
        except json.JSONDecodeError as error:
            raise ApiError(
                400, "VALIDATION_ERROR", "Column mapping must be valid JSON"
            ) from error
        if not isinstance(result, dict):
            raise ApiError(400, "VALIDATION_ERROR", "Column mapping must be an object")
    else:
        result = {key: label for key, label in HEADERS.items() if label in headers}
        result["customFields"] = {}
    for required in ("lastName", "firstName", "birthDate", "personType", "phone"):
        if result.get(required) not in headers:
            raise ApiError(
                400,
                "VALIDATION_ERROR",
                f"Required column is missing: {HEADERS[required]}",
            )
    mapped = [value for key, value in result.items() if key != "customFields" and value]
    mapped.extend((result.get("customFields") or {}).values())
    if len(set(mapped)) != len(mapped) or any(value not in headers for value in mapped):
        raise ApiError(400, "VALIDATION_ERROR", "Column mapping is invalid")
    result.setdefault("customFields", {})
    return result


def _date(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            pass
    return None


def _cell_text(by_header: dict[str, Any], mapping: dict[str, Any], key: str) -> str:
    return str(by_header[mapping[key]].value or "").strip() if mapping.get(key) else ""


def _parse(
    source: bytes, supplied_mapping: str | None = None
) -> tuple[list[str], dict[str, Any], list[dict[str, Any]]]:
    if not source or len(source) > MAX_FILE or not source.startswith(b"PK"):
        raise ApiError(400, "VALIDATION_ERROR", "A valid XLSX up to 5 MiB is required")
    try:
        workbook = load_workbook(io.BytesIO(source), data_only=False, read_only=False)
    except Exception as error:
        raise ApiError(
            400, "VALIDATION_ERROR", "XLSX workbook could not be read"
        ) from error
    sheets = [sheet for sheet in workbook.worksheets if sheet.max_row > 0]
    if len(sheets) != 1 or sheets[0].merged_cells.ranges:
        raise ApiError(
            400, "VALIDATION_ERROR", "Workbook must contain one unmerged worksheet"
        )
    sheet = sheets[0]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if (
        not headers
        or any(not item for item in headers)
        or len(set(item.casefold() for item in headers)) != len(headers)
    ):
        raise ApiError(
            400, "VALIDATION_ERROR", "Header names must be non-empty and unique"
        )
    mapping = _mapping(headers, supplied_mapping)
    parsed: list[dict[str, Any]] = []
    for number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=False), start=2
    ):
        if all(cell.value in (None, "") for cell in values):
            continue
        if len(parsed) >= MAX_ROWS:
            raise ApiError(
                400, "VALIDATION_ERROR", "Workbook contains more than 5000 rows"
            )
        by_header = dict(zip(headers, values, strict=False))
        errors = [
            f"Формула не разрешена: {header}"
            for header, cell in by_header.items()
            if cell.data_type == "f"
        ]

        raw = {
            "lastName": _cell_text(by_header, mapping, "lastName"),
            "firstName": _cell_text(by_header, mapping, "firstName"),
            "middleName": _cell_text(by_header, mapping, "middleName") or None,
            "birthDate": _date(by_header[mapping["birthDate"]].value),
            "personType": _cell_text(by_header, mapping, "personType").upper(),
            "studyGroup": _cell_text(by_header, mapping, "studyGroup") or None,
            "organization": _cell_text(by_header, mapping, "organization") or None,
            "phone": _cell_text(by_header, mapping, "phone") or None,
            "email": _cell_text(by_header, mapping, "email") or None,
        }
        try:
            values_model = ParticipantValues.model_validate(
                {**raw, "customAnswers": []}
            )
        except Exception:
            values_model = None
            errors.append("Данные участника не прошли проверку")
        parsed.append(
            {
                "rowNumber": number,
                "participant": raw,
                "values": values_model,
                "errors": errors,
            }
        )
    if not parsed:
        raise ApiError(400, "VALIDATION_ERROR", "Workbook has no data rows")
    return headers, mapping, parsed


def _event(connection: Any, event_id: str, lock: bool = False) -> Any:
    item = row(
        connection,
        f"SELECT * FROM events WHERE id=:id{' FOR UPDATE' if lock else ''}",
        {"id": event_id},
    )
    if not item:
        raise ApiError(404, "NOT_FOUND", "Event not found")
    if item["status"] == "ARCHIVED":
        raise ApiError(409, "CONFLICT", "Archived Event cannot be imported")
    return item


def _classify(
    connection: Any, event_id: str, parsed: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    result = []
    for item in parsed:
        category, candidates = ("ERROR", []) if item["errors"] else ("NEW", [])
        if not item["errors"]:
            value = item["values"]
            people = rows(
                connection,
                """SELECT p.id,p.last_name,p.first_name,p.middle_name,
                EXISTS(SELECT 1 FROM registrations r WHERE r.person_id=p.id AND r.event_id=:event AND r.status='ACTIVE') active
                FROM persons p WHERE p.merged_into_id IS NULL AND lower(p.last_name)=lower(:last)
                AND lower(p.first_name)=lower(:first) LIMIT 10""",
                {"event": event_id, "last": value.last_name, "first": value.first_name},
            )
            if any(person["active"] for person in people):
                category = "ALREADY_REGISTERED"
            elif people:
                category = "POSSIBLE_MATCH"
                candidates = [
                    {
                        "personId": person["id"],
                        "displayName": " ".join(
                            filter(
                                None,
                                (
                                    person["last_name"],
                                    person["first_name"],
                                    person["middle_name"],
                                ),
                            )
                        ),
                        "matchReason": "PROFILE_SIMILARITY",
                    }
                    for person in people
                ]
        result.append({**item, "category": category, "candidates": candidates})
    return result


@router.post("/{event_id}/import/preview", status_code=201)
async def preview(
    event_id: UUID,
    file: Annotated[UploadFile, File()],
    staff: Annotated[Staff, Depends(csrf_super_admin)],
    db: Annotated[Database, Depends(database)],
    mapping: Annotated[str | None, Form()] = None,
) -> dict[str, Any]:
    source = await file.read(MAX_FILE + 1)
    headers, resolved, parsed = _parse(source, mapping)
    expires = datetime.now(UTC).replace(tzinfo=None) + timedelta(hours=24)
    with db.transaction() as connection:
        event = _event(connection, str(event_id))
        classified = _classify(connection, str(event_id), parsed)
        active_row = row(
            connection,
            "SELECT count(*) count FROM registrations WHERE event_id=:event AND status='ACTIVE'",
            {"event": str(event_id)},
        )
        active = int(active_row["count"] if active_row else 0)

        def count(category: str) -> int:
            return sum(item["category"] == category for item in classified)

        summary = {
            "totalRows": len(classified),
            "newRows": count("NEW"),
            "alreadyRegisteredRows": count("ALREADY_REGISTERED"),
            "possibleMatchRows": count("POSSIBLE_MATCH"),
            "errorRows": count("ERROR"),
            "withoutEmailRows": sum(
                not item["participant"]["email"] for item in classified
            ),
            "capacityImpact": count("NEW") + count("POSSIBLE_MATCH"),
            "activeRegistrations": active,
            "capacity": event["capacity"],
            "exceedsCapacity": active + count("NEW") + count("POSSIBLE_MATCH")
            > event["capacity"],
        }
        job_id = str(uuid4())
        execute(
            connection,
            """INSERT INTO import_jobs (id,event_id,created_by,status,total_rows,valid_rows,error_rows,duplicate_rows,result_summary,expires_at,created_at,updated_at)
            VALUES (:id,:event,:actor,'PREVIEW_READY',:total,:valid,:errors,:duplicates,:summary,:expires,UTC_TIMESTAMP(3),UTC_TIMESTAMP(3))""",
            {
                "id": job_id,
                "event": str(event_id),
                "actor": staff.id,
                "total": len(classified),
                "valid": len(classified) - count("ERROR"),
                "errors": count("ERROR"),
                "duplicates": count("ALREADY_REGISTERED"),
                "summary": db_json(summary),
                "expires": expires,
            },
        )
        execute(
            connection,
            "INSERT INTO import_job_files(import_job_id,file_data,sha256,created_at,expires_at) VALUES (:id,:data,:hash,UTC_TIMESTAMP(3),:expires)",
            {
                "id": job_id,
                "data": source,
                "hash": hashlib.sha256(source).hexdigest(),
                "expires": expires,
            },
        )
    return {
        "importJobId": job_id,
        "expiresAt": expires.replace(tzinfo=UTC).isoformat(),
        "headers": headers,
        "mapping": resolved,
        "summary": summary,
        "rows": [
            {
                key: item[key]
                for key in (
                    "rowNumber",
                    "category",
                    "errors",
                    "participant",
                    "candidates",
                )
            }
            for item in classified
        ],
    }


@router.post("/{event_id}/import/{job_id}/commit")
def commit(
    event_id: UUID,
    job_id: UUID,
    body: dict[str, Any],
    staff: Annotated[Staff, Depends(csrf_super_admin)],
    db: Annotated[Database, Depends(database)],
    config: Annotated[Settings, Depends(settings)],
) -> dict[str, Any]:
    decisions = {int(item["rowNumber"]): item for item in body.get("decisions", [])}
    with db.transaction() as connection:
        job = row(
            connection,
            """SELECT j.*,f.file_data FROM import_jobs j LEFT JOIN import_job_files f ON f.import_job_id=j.id
            WHERE j.id=:id AND j.event_id=:event FOR UPDATE""",
            {"id": str(job_id), "event": str(event_id)},
        )
        if (
            not job
            or job["status"] != "PREVIEW_READY"
            or job["expires_at"] <= datetime.now(UTC).replace(tzinfo=None)
            or not job["file_data"]
        ):
            raise ApiError(409, "CONFLICT", "Import preview is no longer available")
        mapping_json = json.dumps(body.get("mapping", {}), ensure_ascii=False)
        _, _, parsed = _parse(bytes(job["file_data"]), mapping_json)
        classified = _classify(connection, str(event_id), parsed)
        event = _event(connection, str(event_id), True)
        imported = skipped = duplicates = errors = without_email = 0
        for item in classified:
            if item["category"] == "ERROR":
                errors += 1
                continue
            if item["category"] == "ALREADY_REGISTERED":
                duplicates += 1
                continue
            decision = decisions.get(item["rowNumber"])
            if item["category"] == "POSSIBLE_MATCH" and not decision:
                raise ApiError(409, "CONFLICT", "Every possible match needs a decision")
            if decision and decision.get("action") == "SKIP":
                skipped += 1
                continue
            data = participant(item["values"])
            person_id = (
                str(decision["personId"])
                if decision and decision.get("action") == "USE_PERSON"
                else find_or_create_person(connection, data)
            )
            existing = row(
                connection,
                "SELECT id FROM registrations WHERE event_id=:event AND person_id=:person AND status='ACTIVE'",
                {"event": str(event_id), "person": person_id},
            )
            if existing:
                duplicates += 1
                continue
            active_row = row(
                connection,
                "SELECT count(*) count FROM registrations WHERE event_id=:event AND status='ACTIVE'",
                {"event": str(event_id)},
            )
            active = int(active_row["count"] if active_row else 0)
            if active >= int(event["capacity"]) and not body.get(
                "capacityOverride", False
            ):
                raise ApiError(
                    409, "CAPACITY_FULL", "Import would exceed Event capacity"
                )
            create_registration(
                connection,
                str(event_id),
                person_id,
                data,
                "EXCEL_IMPORT",
                False,
                config,
            )
            imported += 1
            without_email += not bool(data["email"])
        result = {
            "importJobId": str(job_id),
            "importedRows": imported,
            "skippedRows": skipped,
            "duplicateRows": duplicates,
            "errorRows": errors,
            "withoutEmailRows": without_email,
        }
        execute(
            connection,
            "UPDATE import_jobs SET status='COMPLETED',result_summary=:summary,committed_at=UTC_TIMESTAMP(3),updated_at=UTC_TIMESTAMP(3) WHERE id=:id",
            {"id": str(job_id), "summary": db_json(result)},
        )
        execute(
            connection,
            "DELETE FROM import_job_files WHERE import_job_id=:id",
            {"id": str(job_id)},
        )
    return result


@router.get("/{event_id}/export.xlsx")
def export(
    event_id: UUID,
    _: Annotated[Staff, Depends(super_admin)],
    db: Annotated[Database, Depends(database)],
) -> StreamingResponse:
    with db.connect() as connection:
        event = _event(connection, str(event_id))
        registrations = rows(
            connection,
            "SELECT * FROM registrations WHERE event_id=:event ORDER BY registered_at,id",
            {"event": str(event_id)},
        )
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Участники"
    columns = [
        "Фамилия",
        "Имя",
        "Отчество",
        "Дата рождения",
        "Тип участника",
        "Группа",
        "Организация",
        "Телефон",
        "Email",
        "Статус",
        "Регистрация",
        "Первое посещение",
    ]
    sheet.append(columns)
    for item in registrations:
        sheet.append(
            [
                _safe(value)
                for value in (
                    item["last_name"],
                    item["first_name"],
                    item["middle_name"],
                    item["birth_date"],
                    item["person_type"],
                    item["study_group"],
                    item["organization"],
                    item["phone"],
                    item["email"],
                    item["status"],
                    item["registered_at"],
                    item["first_attended_at"],
                )
            ]
        )
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"event-{event['slug']}-participants.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "private, no-store",
        },
    )
