from __future__ import annotations

from io import BytesIO
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

from event_api.config import Settings
from event_api.database import Database
from event_api.demo_seed import main as seed_demo
from event_api.email_worker import process_once
from event_api.errors import ApiError
from event_api.routers.excel import _parse
from event_api.security import RateLimiter, auth_link_token, hash_password, token_hash

ORIGIN = {"Origin": "http://localhost:5173"}


def test_excel_rejects_formula_and_merged_cells() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["Фамилия", "Имя", "Дата рождения", "Тип участника", "Телефон"])
    sheet.append(["=1+1", "Иван", "2000-01-01", "KAIT_TEACHER", "+79990000000"])
    source = BytesIO()
    workbook.save(source)
    assert _parse(source.getvalue())[2][0]["errors"]
    sheet.merge_cells("A2:B2")
    source = BytesIO()
    workbook.save(source)
    with pytest.raises(ApiError):
        _parse(source.getvalue())


def _seed_admin(client: TestClient) -> str:
    database: Database = client.app.state.database
    user_id = str(uuid4())
    with database.transaction() as connection:
        connection.execute(
            text("""INSERT INTO staff_users
            (id,email,email_normalized,password_hash,system_role,active,password_changed_at,created_at,updated_at)
            VALUES (:id,'admin@example.com','admin@example.com',:password,'SUPER_ADMIN',TRUE,UTC_TIMESTAMP(3),UTC_TIMESTAMP(3),UTC_TIMESTAMP(3))"""),
            {"id": user_id, "password": hash_password("correct horse battery")},
        )
    return user_id


def _login(client: TestClient) -> tuple[dict[str, str], str]:
    client.cookies.clear()
    response = client.post(
        "/auth/login",
        headers=ORIGIN,
        json={"email": "admin@example.com", "password": "correct horse battery"},
    )
    assert response.status_code == 200, response.text
    csrf = response.json()["csrfToken"]
    return {**ORIGIN, "X-CSRF-Token": csrf}, response.cookies["staff_session"]


def test_health_and_security_foundation(client: TestClient) -> None:
    health = client.get("/health/ready")
    assert health.json() == {"status": "ready"}
    assert health.headers["x-content-type-options"] == "nosniff"
    assert health.headers["x-frame-options"] == "DENY"
    _seed_admin(client)
    unknown = client.post(
        "/auth/login",
        headers=ORIGIN,
        json={"email": "missing@example.com", "password": "incorrect password"},
    )
    invalid = client.post(
        "/auth/login",
        headers=ORIGIN,
        json={"email": "admin@example.com", "password": "incorrect password"},
    )
    assert (unknown.status_code, unknown.json()["error"]["code"]) == (
        401,
        "INVALID_CREDENTIALS",
    )
    assert invalid.json()["error"]["code"] == "INVALID_CREDENTIALS"
    client.cookies.set("staff_session", "revoked-or-malformed-cookie")
    assert (
        client.post(
            "/auth/login",
            headers=ORIGIN,
            json={"email": "admin@example.com", "password": "correct horse battery"},
        ).status_code
        == 200
    )
    headers, raw_session = _login(client)
    database: Database = client.app.state.database
    with database.connect() as connection:
        stored = connection.execute(
            text("SELECT token_hash FROM sessions WHERE token_hash=:hash"),
            {"hash": token_hash(raw_session)},
        ).scalar_one()
    assert stored == token_hash(raw_session)
    assert raw_session != stored
    assert (
        client.post(
            "/admin/events", json={}, headers={"Origin": "https://evil.test"}
        ).status_code
        == 403
    )
    assert client.post("/auth/logout", headers=ORIGIN).status_code == 403
    assert client.post("/auth/logout", headers=headers).status_code == 200
    assert client.get("/auth/session").status_code == 401


def test_production_configuration_hides_api_schema(database_url: str) -> None:
    from event_api.main import create_app

    config = Settings(
        database_url=database_url,
        node_env="production",
        cors_origins=["https://events.example.org", "https://scanner.example.org"],
        session_secret="s" * 43,
        auth_link_secret="a" * 43,
        auth_link_base_url="https://events.example.org/auth",
        qr_signing_secret="q" * 43,
        public_web_base_url="https://events.example.org",
        consent_url="https://example.org/privacy",
        consent_version="2026-08-26",
    )
    with pytest.raises(ValueError, match="STARTTLS"):
        Settings.model_validate(
            {
                **config.model_dump(),
                "smtp_host": "smtp.example.org",
                "smtp_from_email": "noreply@example.org",
                "smtp_starttls": False,
            }
        )
    with TestClient(create_app(config)) as production:
        assert production.get("/docs").status_code == 404
        assert production.get("/openapi.json").status_code == 404
        response = production.get("/health")
        assert "max-age=31536000" in response.headers["strict-transport-security"]


def test_shared_rate_limiter_hides_identifiers(client: TestClient) -> None:
    database: Database = client.app.state.database
    config = client.app.state.settings.model_copy(
        update={"auth_rate_limit_max": 2, "auth_rate_limit_window_seconds": 60}
    )
    limiter = RateLimiter(config, database)
    limiter.consume("test-release", "person@example.org")
    limiter.consume("test-release", "person@example.org")
    with pytest.raises(ApiError) as caught:
        limiter.consume("test-release", "person@example.org")
    assert caught.value.status == 429
    with database.connect() as connection:
        keys = connection.execute(
            text("SELECT bucket_key FROM security_rate_limits")
        ).scalars()
    assert all("person@example.org" not in key for key in keys)


def test_domain_constraints_and_event_crud(client: TestClient) -> None:
    headers, _ = _login(client)
    payload = {
        "title": "Python MVP",
        "slug": "python-mvp",
        "description": "Integration event",
        "startAt": "2026-10-10T10:00:00Z",
        "endAt": "2026-10-10T12:00:00Z",
        "timezone": "Europe/Moscow",
        "location": "КАИТ №20",
        "registrationDeadline": "2026-10-09T10:00:00Z",
        "capacity": 100,
        "status": "DRAFT",
    }
    created = client.post("/admin/events", headers=headers, json=payload)
    assert created.status_code == 201, created.text
    event_id = created.json()["id"]
    assert client.get(f"/admin/events/{event_id}").status_code == 200
    field = client.post(
        f"/admin/events/{event_id}/form-fields",
        headers=headers,
        json={"type": "BOOLEAN", "label": "Согласие", "required": True, "sortOrder": 1},
    )
    assert field.status_code == 201, field.text
    assert (
        client.delete(
            f"/admin/events/{event_id}/form-fields/{field.json()['id']}",
            headers=headers,
        ).status_code
        == 200
    )


def test_mysql_specific_invariants(client: TestClient) -> None:
    database: Database = client.app.state.database
    with database.connect() as connection:
        version = str(connection.execute(text("SELECT VERSION()")).scalar_one())
        tables = {row[0] for row in connection.execute(text("SHOW TABLES"))}
        columns = {
            row[0] for row in connection.execute(text("SHOW COLUMNS FROM sessions"))
        }
    assert version.startswith("8.1.0")
    assert {
        "persons",
        "events",
        "registrations",
        "attendance_events",
        "schema_migrations",
    } <= tables
    assert "token_hash" in columns and "token" not in columns


def test_demo_seed_is_idempotent(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    from event_api.config import get_settings

    monkeypatch.setenv("NODE_ENV", "development")
    monkeypatch.setenv("DEMO_ADMIN_EMAIL", "demo-admin@example.com")
    monkeypatch.setenv("DEMO_ADMIN_PASSWORD", "demo admin safe password")
    monkeypatch.setenv("DEMO_SCANNER_EMAIL", "demo-scanner@example.com")
    monkeypatch.setenv("DEMO_SCANNER_PASSWORD", "demo scanner safe password")
    get_settings.cache_clear()
    try:
        seed_demo()
        seed_demo()
    finally:
        get_settings.cache_clear()
    database: Database = client.app.state.database
    with database.connect() as connection:
        event_count = connection.execute(
            text("SELECT count(*) FROM events WHERE slug='demo-event'")
        ).scalar_one()
        access_count = connection.execute(
            text("""SELECT count(*) FROM event_access a
            JOIN events e ON e.id=a.event_id JOIN staff_users u ON u.id=a.user_id
            WHERE e.slug='demo-event' AND u.email_normalized='demo-scanner@example.com'""")
        ).scalar_one()
    assert event_count == 1 and access_count == 1


def test_public_registration_and_idempotent_attendance(client: TestClient) -> None:
    headers, _ = _login(client)
    database: Database = client.app.state.database
    with database.connect() as connection:
        event_id = connection.execute(
            text("SELECT id FROM events WHERE slug='python-mvp'")
        ).scalar_one()
    opened = client.patch(
        f"/admin/events/{event_id}",
        headers=headers,
        json={"status": "REGISTRATION_OPEN"},
    )
    assert opened.status_code == 200, opened.text
    client.cookies.clear()
    registration = client.post(
        "/public/events/python-mvp/register",
        headers=ORIGIN,
        json={
            "lastName": "Иванов",
            "firstName": "Иван",
            "middleName": "Иванович",
            "birthDate": "2005-03-15",
            "email": "participant@example.com",
            "phone": "+79991234567",
            "studyGroup": "ИС-21",
            "personType": "KAIT_STUDENT",
            "customAnswers": [],
            "consentAccepted": True,
            "consentVersion": "test-v1",
        },
    )
    assert registration.status_code == 201, registration.text
    repeated = client.post(
        "/public/events/python-mvp/register",
        headers=ORIGIN,
        json={
            "lastName": "Иванов",
            "firstName": "Иван",
            "middleName": "Иванович",
            "birthDate": "2005-03-15",
            "email": "participant@example.com",
            "phone": "+79991234567",
            "studyGroup": "ИС-21",
            "personType": "KAIT_STUDENT",
            "customAnswers": [],
            "consentAccepted": True,
            "consentVersion": "test-v1",
        },
    )
    assert repeated.status_code == 200
    with database.connect() as connection:
        registration_id = connection.execute(
            text(
                "SELECT id FROM registrations WHERE event_id=:event AND status='ACTIVE'"
            ),
            {"event": event_id},
        ).scalar_one()
    headers, _ = _login(client)
    client_event_id = str(uuid4())
    attendance = {
        "deviceId": str(uuid4()),
        "events": [
            {
                "clientEventId": client_event_id,
                "registrationId": registration_id,
                "mode": "MANUAL_CONFIRM",
                "source": "ONLINE",
                "deviceScannedAt": "2026-10-10T10:30:00Z",
                "estimatedScannedAt": "2026-10-10T10:30:00Z",
            }
        ],
    }
    accepted = client.post(
        f"/scanner/events/{event_id}/attendance/sync",
        headers=headers,
        json=attendance,
    )
    assert accepted.status_code == 201, accepted.text
    assert accepted.json()["results"][0]["status"] == "ACCEPTED"
    retried = client.post(
        f"/scanner/events/{event_id}/attendance/sync",
        headers=headers,
        json=attendance,
    )
    assert retried.json()["results"][0]["status"] == "ALREADY_PROCESSED"


def test_scanner_invitation_and_event_access(client: TestClient) -> None:
    headers, _ = _login(client)
    database: Database = client.app.state.database
    with database.connect() as connection:
        event_id = connection.execute(
            text("SELECT id FROM events WHERE slug='python-mvp'")
        ).scalar_one()
    invited = client.post(
        "/admin/staff/invitations",
        headers=headers,
        json={"email": "scanner@example.com", "eventId": event_id},
    )
    assert invited.status_code == 201, invited.text
    invitation_id = invited.json()["id"]
    with database.connect() as connection:
        record = (
            connection.execute(
                text(
                    "SELECT expires_at,token_hash FROM staff_invitations WHERE id=:id"
                ),
                {"id": invitation_id},
            )
            .mappings()
            .one()
        )
    token = auth_link_token(
        "invitation",
        invitation_id,
        record["expires_at"],
        client.app.state.settings.auth_link_secret,
    )
    assert token_hash(token) == record["token_hash"] and token != record["token_hash"]
    client.cookies.clear()
    accepted = client.post(
        f"/auth/invitations/{token}/accept",
        headers=ORIGIN,
        json={"password": "scanner secure password"},
    )
    assert accepted.status_code == 200, accepted.text
    assert (
        client.post(
            f"/auth/invitations/{token}/accept",
            headers=ORIGIN,
            json={"password": "scanner secure password"},
        ).status_code
        == 400
    )
    scanner_login = client.post(
        "/auth/login",
        headers=ORIGIN,
        json={"email": "scanner@example.com", "password": "scanner secure password"},
    )
    assert scanner_login.status_code == 200
    scanner_headers = {**ORIGIN, "X-CSRF-Token": scanner_login.json()["csrfToken"]}
    visible = client.get("/scanner/events")
    assert visible.status_code == 200
    assert any(item["id"] == event_id for item in visible.json()["items"])
    assert (
        client.post("/admin/events", headers=scanner_headers, json={}).status_code
        == 403
    )


def test_excel_preview_commit_and_safe_export(client: TestClient) -> None:
    headers, _ = _login(client)
    database: Database = client.app.state.database
    with database.connect() as connection:
        event_id = connection.execute(
            text("SELECT id FROM events WHERE slug='python-mvp'")
        ).scalar_one()
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(
        [
            "Фамилия",
            "Имя",
            "Отчество",
            "Дата рождения",
            "Тип участника",
            "Группа",
            "Организация",
            "Телефон",
            "Email",
        ]
    )
    sheet.append(
        [
            "Петрова",
            "Анна",
            "Сергеевна",
            "1990-01-10",
            "KAIT_TEACHER",
            "",
            "КАИТ №20",
            "+79997654321",
            "excel@example.com",
        ]
    )
    source = BytesIO()
    workbook.save(source)
    preview = client.post(
        f"/admin/events/{event_id}/import/preview",
        headers=headers,
        files={
            "file": (
                "participants.xlsx",
                source.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 201, preview.text
    assert preview.json()["summary"]["newRows"] == 1
    commit = client.post(
        f"/admin/events/{event_id}/import/{preview.json()['importJobId']}/commit",
        headers=headers,
        json={
            "mapping": preview.json()["mapping"],
            "decisions": [],
            "capacityOverride": False,
        },
    )
    assert commit.status_code == 200, commit.text
    assert commit.json()["importedRows"] == 1
    repeated = client.post(
        f"/admin/events/{event_id}/import/{preview.json()['importJobId']}/commit",
        headers=headers,
        json={"mapping": preview.json()["mapping"], "decisions": []},
    )
    assert repeated.status_code == 409
    exported = client.get(f"/admin/events/{event_id}/export.xlsx")
    assert exported.status_code == 200
    assert exported.content.startswith(b"PK")
    with database.connect() as connection:
        assert (
            connection.execute(
                text("SELECT count(*) FROM import_job_files")
            ).scalar_one()
            == 0
        )


def test_password_reset_is_one_time_and_revokes_sessions(client: TestClient) -> None:
    headers, raw_session = _login(client)
    forgot = client.post(
        "/auth/password/forgot",
        headers=headers,
        json={"email": "admin@example.com"},
    )
    assert forgot.status_code == 202
    database: Database = client.app.state.database
    with database.connect() as connection:
        record = (
            connection.execute(
                text(
                    "SELECT id,expires_at,token_hash FROM password_reset_tokens ORDER BY created_at DESC LIMIT 1"
                )
            )
            .mappings()
            .one()
        )
    token = auth_link_token(
        "password-reset",
        record["id"],
        record["expires_at"],
        client.app.state.settings.auth_link_secret,
    )
    assert token_hash(token) == record["token_hash"] and token != record["token_hash"]
    reset = client.post(
        "/auth/password/reset",
        headers=headers,
        json={"token": token, "password": "new correct horse battery"},
    )
    assert reset.status_code == 200, reset.text
    assert client.get("/auth/session").status_code == 401
    assert (
        client.post(
            "/auth/password/reset",
            headers=headers,
            json={"token": token, "password": "another correct password"},
        ).status_code
        == 400
    )
    with database.connect() as connection:
        revoked = connection.execute(
            text("SELECT revoked_at FROM sessions WHERE token_hash=:hash"),
            {"hash": token_hash(raw_session)},
        ).scalar_one()
    assert revoked is not None


def test_email_worker_sends_durable_intent_without_persisting_link(
    client: TestClient,
) -> None:
    database: Database = client.app.state.database
    config = client.app.state.settings.model_copy(
        update={
            "smtp_host": "smtp.example.test",
            "smtp_from_email": "noreply@example.test",
        }
    )
    captured: list[str] = []

    def fake_sender(message: object, _config: Settings) -> str:
        captured.append(str(message))
        return "provider-test-id"

    assert process_once(database, config, fake_sender) == 1
    assert captured and "provider-test-id" not in captured[0]
    with database.connect() as connection:
        delivery = (
            connection.execute(
                text(
                    "SELECT status,provider_message_id FROM email_deliveries WHERE status='SENT' ORDER BY sent_at DESC LIMIT 1"
                )
            )
            .mappings()
            .one()
        )
    assert delivery == {"status": "SENT", "provider_message_id": "provider-test-id"}
